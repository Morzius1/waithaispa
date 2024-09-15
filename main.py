from ast import main
from bs4 import BeautifulSoup
import requests
from fake_useragent import UserAgent
import random
import lxml
import re 
from seleniumbase import SB
from seleniumbase.common.exceptions import ElementNotVisibleException,NoSuchElementException
import time
from openpyxl import Workbook,load_workbook


class Work:

    def __init__(self,proxies:dict,headers:dict) -> None:
        self.proxies=proxies
        self.headers=headers
        # self.failed_salons_urls=[]#Формируется при применении функции get_dict_salon_id_and_masters_id, так как не на всех сайтах имеется id салона и id мастеров данного салона
        self.failed_salons_urls=['https://waithaispa.ru/contacts/hodinka.html',
                                 'https://waithaispa.ru/contacts/li%D0%B0nozovo.html',
                                 'https://waithaispa.ru/contacts/mayakovskaya.html',
                                 'https://waithaispa.ru/contacts/mayakovskaya2.html',
                                 'https://waithaispa.ru/contacts/riga.html']#Формируется при применении функции get_dict_salon_id_and_masters_id, так как не на всех сайтах имеется id салона и id мастеров данного салона

    #Данной функцией получаем url всех салонов Москвы и МО
    def get_urls(self):
        response=requests.get('https://waithaispa.ru/contacts/',proxies=self.proxies,headers=self.headers)
        soup=BeautifulSoup(response.text,'lxml')
        result=soup.find_all('div',class_=re.compile('contacts__balloon balloon stocks-card moscow_mo'))
        for i in result:
            url=f'https://waithaispa.ru{i.find('a',{'itemprop':'name'}).get('href')}'
            print(url)
            yield url

    """Данной функцией собираем id салонов и мастеров сайтом, причем не каждый сайт содержит данные id.
    Такие ID будут помещаться в список self.failed_salons_urls и обрабатываться отдельной функцией в дальнейшем (name function)"""
    def get_dict_salon_id_and_masters_id(self):
        for i in self.get_urls():
            try:
                response=requests.get(i,proxies=self.proxies,headers=self.headers)
                soup=BeautifulSoup(response.text,'lxml')
                res=soup.find('iframe',class_='widget-wt-frame').get('src') 
                salon_id=res.split('&')[1].split('=')[1]
                masters_id=res.split('&')[2].split('=')[1]
                print(salon_id,masters_id,i)
                time.sleep(3)
                yield salon_id,masters_id,i
            except AttributeError as ae:
                print(f'Произошла ошибка: {i}',ae)
                self.failed_salons_urls.append(i)
            except IndexError as ie:
                print(f'Произошла ошибка: {i}',ie)
                self.failed_salons_urls.append(i)
            except TypeError as te:
                print(f'Произошла ошибка: {i}',te)
                self.failed_salons_urls.append(i)

    """Данной функцией реализуем получение информации о салонах и мастерах, которые работают в данных салонах соответственно"""
    def get_info_about_salons_and_masters(self):
        #Сначала соберем информацию о салоне и затем сразу о мастерах, чтобы дважды не вызывать метод с получением ID с каждого url
        for k in self.get_dict_salon_id_and_masters_id():
            response=requests.get(f'https://widget-api.waithaispa.ru/api/get_settings/?id={k[0]}',proxies=self.proxies,headers=self.headers)
            res=response.json()
            name_salon=res['name']
            address=res['geo_address'],
            phone=*res['phones'],
            url_salon=k[2],
            url_call=f"https://booking-promo.waithaispa.ru/?new_widget=1&SalonId={k[0]}"
            #Собираем информацию о мастерах
            response_masters=requests.get(f'https://waithaispa.ru/si_nocache/master.php?id={k[1]}',proxies=self.proxies,headers=self.headers)
            soup=BeautifulSoup(response_masters.text,'lxml')
            info=soup.find_all('div',class_='s-masters-modal')
            for i in info:
                try:
                    master_name=i.find('div',class_='s-masters-modal__name').text.strip()
                    master_desc=i.find('div',class_='salon-list-mastera-item__text').text.strip()
                    url_photo=f"https://waithaispa.ru{i.find('img',class_='s-masters-modal__img').get("src")}"
                    print(name_salon,*address,*phone,*url_salon,url_call,master_name,master_desc,url_photo)
                    yield [name_salon,*address,*phone,*url_salon,url_call,master_name,master_desc,url_photo]
                except AttributeError as ex:
                    print(ex,'Мастер не имеет описания')
                    master_name=i.find('div',class_='s-masters-modal__name').text
                    master_desc=''
                    url_photo=f"https://waithaispa.ru{i.find('img',class_='s-masters-modal__img').get("src")}"
                    print(name_salon,*address,*phone,*url_salon,url_call,master_name,master_desc,url_photo)
                    yield [name_salon,*address,*phone,*url_salon,url_call,master_name,master_desc,url_photo]
                time.sleep(3)

    def data_to_excel(self,output_name):
        wb=Workbook()
        ws=wb.active
        ws.title='inf'
        ws.append(['Наименование',"Адрес","Номер(а) телефона(ов)","Ссылка на сайт","Ссылка на запись","Имя мастера","Описание мастера","Ссылка на фото"])

        # Получение информации c cайтов на которых есть ID салона и мастеров
        for j in self.get_info_about_salons_and_masters():
            ws.append(j)
        wb.save(f'{output_name}.xlsx')
        # Сохранение для того что если мы получим ошибку в дальнейшем коде, остальные данные у нас останутся 
        # Получение информации c cайтов на которых нет ID салона и мастеров
        wb = load_workbook(f'{output_name}.xlsx')
        ws = wb.active
        wb=Workbook()
        ws=wb.active
        ws.title='inf'
        ws.append(['Наименование',"Адрес","Номер(а) телефона(ов)","Ссылка на сайт","Ссылка на запись","Имя мастера","Описание мастера","Ссылка на фото"])
        with SB(sjw=True,pls='none',uc=True) as sb:
            for i in self.failed_salons_urls:
                sb.open(i)
                time.sleep(10)
                name_salon=sb.find_element('span[class="bredcrumbrs__link color-white"]').text
                address_salon=sb.find_element('span[itemprop="streetAddress"]')
                phone_number=sb.find_elements('a[itemprop="telephone"]')
                phones_number=[]
                for k in phone_number:
                    phones_number.append(k.text)
                phones_number=' '.join(phones_number)
                urls=sb.find_element('div[class="swiper-slide swiper-slide-active"]')
                page_scroll=2
                amount_of_masters=int(urls.get_attribute('aria-label')[-1])
                for j in range(amount_of_masters):
                    try:
                        sb.click(f'div[data-path="master_{j}"]')
                        master_name=sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__master > div.s-masters-modal__info > div').text.strip()
                        master_desc=sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__desc > div').text.strip()
                        url_photo=f"{sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__master > div.s-masters-modal__box-img > img').get_attribute("src")}"
                        ws.append([name_salon,address_salon.text,phones_number,i,'Новый салон, запись только через сайт самого салона',master_name,master_desc,url_photo])
                        time.sleep(2)
                        sb.click('#s-masters__block > div.modal.is-open > div > button > svg > use')
                        time.sleep(3)
                    except ElementNotVisibleException as ex:
                        print(ex)
                        if amount_of_masters>4:
                            try:
                                sb.click(f"#salon_master > div.container > div > div.s-masters__pagination.swiper-pagination.swiper-pagination-clickable.swiper-pagination-bullets.swiper-pagination-horizontal > span:nth-child({page_scroll})")
                                page_scroll+=1
                                sb.click(f'div[data-path="master_{j}"]')
                                master_name=sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__master > div.s-masters-modal__info > div').text.strip()
                                master_desc=sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__desc > div').text.strip()
                                url_photo=f"{sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__master > div.s-masters-modal__box-img > img').get_attribute("src")}"
                                ws.append([name_salon,address_salon.text,phones_number,i,'Новый салон, запись только через сайт самого салона',master_name,master_desc,url_photo])
                                time.sleep(2)
                                sb.click('#s-masters__block > div.modal.is-open > div > button > svg > use')
                                time.sleep(3)
                            except ElementNotVisibleException as eee:
                                try:
                                    sb.click('#salon_master > div.container > div > div.s-masters__pagination.swiper-pagination.swiper-pagination-clickable.swiper-pagination-bullets.swiper-pagination-horizontal > span.swiper-pagination-bullet.swiper-pagination-bullet-active')
                                    page_scroll+=1
                                    sb.click(f'div[data-path="master_{j}"]')
                                    master_name=sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__master > div.s-masters-modal__info > div').text.strip()
                                    master_desc=sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__desc > div').text.strip()
                                    url_photo=f"{sb.find_element('#s-masters__block > div.modal.is-open > div > div > div > div.s-masters-modal__master > div.s-masters-modal__box-img > img').get_attribute("src")}"
                                    ws.append([name_salon,address_salon.text,phones_number,i,'Новый салон, запись только через сайт самого салона',master_name,master_desc,url_photo])
                                    time.sleep(2)
                                    sb.click('#s-masters__block > div.modal.is-open > div > button > svg > use')
                                except NoSuchElementException as r:
                                    print(r,'Элемент не был найден')
                                    continue
                            except NoSuchElementException as r:
                                    print(r,'Элемент не был найден')
                                    continue
                        else:
                                continue
                    except NoSuchElementException as r:
                        print(r,'Элемент не был найден')
                        sb.click('#s-masters__block > div.modal.is-open > div > button > svg > use')
                        continue    

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 4)
            ws.column_dimensions[column].width = adjusted_width
        wb.save(f'{output_name}.xlsx')

def main():
    #Настройка 
    ua=UserAgent()
    headers={
        'User-Agent':ua.random
            }
    proxy_list=['socks5://login:password@ip:port']
    prox_ip=random.choice(proxy_list)
    prox={
            'http': f'{prox_ip}',
            'https':f'{prox_ip}'
            }
    work=Work(prox,headers)
    #Ваше наименование файла на выходе
    work.data_to_excel(output_name='ES')

if __name__=='__main__':
    main()
